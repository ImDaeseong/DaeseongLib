from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook


def loadurlTOTAL():
    responses = requests.get("http://music.naver.com/listen/top100.nhn?domain=TOTAL")
    soup = BeautifulSoup(responses.text, 'lxml')

    titles = soup.select('a._title span.ellipsis')
    artists = soup.select('a._artist span.ellipsis')
    # print(artists)

    titleslist = [title.get_text() for title in titles]
    artistslist = [artist.get_text().strip() for artist in artists]
    # print(len(titleslist))

    if len(titleslist) == len(artistslist):
        for item in range(len(artistslist)):
            print("{0}: {1}".format(titleslist[item], artistslist[item]))


def loadurlDOMESTIC():
    responses = requests.get("http://music.naver.com/listen/top100.nhn?domain=DOMESTIC")
    soup = BeautifulSoup(responses.text, 'lxml')

    titles = soup.select('a._title span.ellipsis')
    artists = soup.select('a._artist span.ellipsis')
    # print(artists)

    titleslist = [title.get_text() for title in titles]
    artistslist = [artist.get_text().strip() for artist in artists]
    # print(len(titleslist))

    if len(titleslist) == len(artistslist):
        for item in range(len(artistslist)):
            print("{0}: {1}".format(titleslist[item], artistslist[item]))


def loadurlOVERSEA():
    responses = requests.get("http://music.naver.com/listen/top100.nhn?domain=OVERSEA")
    soup = BeautifulSoup(responses.text, 'lxml')

    titles = soup.select('a._title span.ellipsis')
    artists = soup.select('a._artist span.ellipsis')
    # print(titles)
    # print(artists)

    titleslist = [title.get_text() for title in titles]
    artistslist = [artist.get_text().strip() for artist in artists]
    # print(titleslist)
    # print(artistslist)
    # print(len(artistslist))

    if len(titleslist) == len(artistslist):
        for item in range(len(artistslist)):
            print("{0}: {1}".format(titleslist[item], artistslist[item]))


def loadurl():
    responses = requests.get("http://music.naver.com/listen/history/index.nhn")
    soup = BeautifulSoup(responses.text, 'lxml')

    titles = soup.select('a._title span.ellipsis')
    artists = soup.select('a._artist span.ellipsis')
    # print(titles)

    titleslist = [title.get_text() for title in titles]
    artistslist = [artist.get_text().strip() for artist in artists]
    # print(titleslist)
    # print(artistslist)
    # print(len(artistslist))

    wb = Workbook()
    ws = wb.create_sheet("naver music", 0)

    # 둘째줄부터 데이터 입력
    row = 2
    # 첫줄은 컬럼 제목 설정
    ws.cell(1, 1, "제목")
    ws.cell(1, 2, "가수")

    if len(titleslist) == len(artistslist):

        for item in range(len(artistslist)):
            print("{0}: {1}".format(titleslist[item], artistslist[item]))
            ws.cell(row, 1, titleslist[item])
            ws.cell(row, 2, artistslist[item])
            row += 1

    wb.save("naver music.xlsx")


if __name__ == "__main__":
    # loadurlTOTAL()
    # loadurlDOMESTIC()
    # loadurlOVERSEA()
    loadurl()
